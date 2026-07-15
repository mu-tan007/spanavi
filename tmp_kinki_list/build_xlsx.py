# -*- coding: utf-8 -*-
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

KAIGO = r'C:\Users\篠宮拓武\.claude\projects\C--Users-----\fd530b4d-5b0a-4e8d-b6c2-37c4b1bf8ab2\tool-results\mcp-plugin_supabase_supabase-execute_sql-1783309532175.txt'
KENSETSU = r'C:\Users\篠宮拓武\.claude\projects\C--Users-----\fd530b4d-5b0a-4e8d-b6c2-37c4b1bf8ab2\tool-results\mcp-plugin_supabase_supabase-execute_sql-1783309710591.txt'
OUTDIR = r'C:\Users\篠宮拓武\Projects\spanavi\tmp_kinki_list'

def parse(path):
    raw = open(path, 'rb').read().decode('utf-8')
    obj = json.loads(raw)
    res = obj['result']
    a = res.index('['); b = res.rindex(']') + 1
    arr = json.loads(res[a:b])          # [{"j": "<jsonstring>"}]
    return json.loads(arr[0]['j'])      # list of row dicts

# (key, header, is_number)
COLS = [
    ('company_name',        '企業名',            False),
    ('tsr_id',              'TSR-ID',            False),
    ('prefecture',          '都道府県',          False),
    ('city',                '市区町村',          False),
    ('address',             '住所',              False),
    ('phone',               '電話番号',          False),
    ('industry_major',      '業種（大分類）',    False),
    ('industry_sub',        '業種（細分類）',    False),
    ('business_description', '事業内容',          False),
    ('revenue_k',           '売上高（千円）',    True),
    ('net_income_k',        '当期純利益（千円）', True),
    ('ordinary_income_k',   '経常利益（千円）',  True),
    ('capital_k',           '資本金（千円）',    True),
    ('employee_count',      '従業員数',          True),
    ('established_year',    '設立年',            False),
    ('representative',      '代表者',            False),
    ('representative_age',  '代表者年齢',        True),
    ('shareholders',        '株主',              False),
    ('officers',            '役員',              False),
]
WIDTHS = [28,11,9,12,30,15,16,16,40,14,16,14,14,9,7,14,9,30,34]

HDR_FILL = PatternFill('solid', fgColor='1F3864')
HDR_FONT = Font(bold=True, color='FFFFFF')

def add_sheet(wb, title, rows):
    ws = wb.active if wb.worksheets and wb.active.title == 'Sheet' and ws_unused(wb) else wb.create_sheet(title)
    ws.title = title
    for c,(key,hdr,isnum) in enumerate(COLS, start=1):
        cell = ws.cell(row=1, column=c, value=hdr)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = WIDTHS[c-1]
    for r,row in enumerate(rows, start=2):
        for c,(key,hdr,isnum) in enumerate(COLS, start=1):
            v = row.get(key)
            cell = ws.cell(row=r, column=c, value=v)
            if isnum and v is not None:
                cell.number_format = '#,##0'
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = '%s1:%s%d' % (get_column_letter(1), get_column_letter(len(COLS)), len(rows)+1)
    return ws

def ws_unused(wb):
    return True

def build(path, sheets):
    wb = Workbook()
    default = wb.active
    for i,(title, rows) in enumerate(sheets):
        if i == 0:
            ws = default; ws.title = title
            for c,(key,hdr,isnum) in enumerate(COLS, start=1):
                cell = ws.cell(row=1, column=c, value=hdr)
                cell.fill = HDR_FILL; cell.font = HDR_FONT
                cell.alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
                ws.column_dimensions[get_column_letter(c)].width = WIDTHS[c-1]
            for r,row in enumerate(rows, start=2):
                for c,(key,hdr,isnum) in enumerate(COLS, start=1):
                    v = row.get(key); cell = ws.cell(row=r, column=c, value=v)
                    if isnum and v is not None: cell.number_format = '#,##0'
            ws.freeze_panes = 'A2'
            ws.auto_filter.ref = 'A1:%s%d' % (get_column_letter(len(COLS)), len(rows)+1)
        else:
            ws = wb.create_sheet(title)
            for c,(key,hdr,isnum) in enumerate(COLS, start=1):
                cell = ws.cell(row=1, column=c, value=hdr)
                cell.fill = HDR_FILL; cell.font = HDR_FONT
                cell.alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
                ws.column_dimensions[get_column_letter(c)].width = WIDTHS[c-1]
            for r,row in enumerate(rows, start=2):
                for c,(key,hdr,isnum) in enumerate(COLS, start=1):
                    v = row.get(key); cell = ws.cell(row=r, column=c, value=v)
                    if isnum and v is not None: cell.number_format = '#,##0'
            ws.freeze_panes = 'A2'
            ws.auto_filter.ref = 'A1:%s%d' % (get_column_letter(len(COLS)), len(rows)+1)
    wb.save(path)

kaigo = parse(KAIGO)
kensetsu = parse(KENSETSU)
kaigo_strict = [r for r in kaigo if r.get('is_strict')]

print('kaigo total', len(kaigo), 'kaigo strict', len(kaigo_strict), 'kensetsu', len(kensetsu))

f1 = OUTDIR + r'\近畿_介護建設リスト_厳格版_20260706.xlsx'
f2 = OUTDIR + r'\近畿_介護建設リスト_介護緩和版_20260706.xlsx'
build(f1, [('建設', kensetsu), ('介護', kaigo_strict)])
build(f2, [('建設', kensetsu), ('介護', kaigo)])
print('saved', f1)
print('saved', f2)
